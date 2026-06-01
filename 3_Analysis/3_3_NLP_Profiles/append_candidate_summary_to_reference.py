# ==========================================================
# append_candidate_summary_to_reference.py
#
# 目的：
# 1. 把 deepseek_candidates_review_by_symptom.xlsx 里
#    “每个症状候选汇总”sheet 的候选表述，
#    去掉 (n=..., variant/high...) 这些统计信息，
#    只保留纯文本表述。
#
# 2. 将这些纯文本表述追加到
#    symptom_expression_reference.xlsx 对应症状行的 E列以后。
#
# 3. 同时导出删除词典 txt：
#
#    delete_for_M4_augmented_pure.txt
#      = C列包含“文本症状”的症状行，E列以后全部表述
#
#    delete_for_M3_augmented_pure.txt
#      = C列包含“文本症状”的症状行
#        +
#        C列包含“问卷症状”且D列为空白的症状行
#        对应 E列以后全部表述
#
#    delete_for_M5_augmented_pure.txt
#      = 所有症状行 E列以后全部表述
#
# 输入：
#   1. symptom_expression_reference.xlsx
#   2. deepseek_symptom_expression_outputs/deepseek_candidates_review_by_symptom.xlsx
#
# 输出：
#   1. symptom_expression_reference_with_deepseek_candidates.xlsx
#   2. deepseek_symptom_expression_outputs/append_candidate_summary_log.xlsx
#   3. delete_for_M3_augmented_pure.txt
#   4. delete_for_M4_augmented_pure.txt
#   5. delete_for_M5_augmented_pure.txt
#
# 注意：
#   不覆盖原始 symptom_expression_reference.xlsx
# ==========================================================

import os
import re
from copy import copy

import pandas as pd
from openpyxl import load_workbook


# ==========================================================
# 0. 路径设置
# ==========================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

REFERENCE_FILE = os.path.join(
    BASE_DIR,
    "symptom_expression_reference.xlsx"
)

CANDIDATE_FILE = os.path.join(
    BASE_DIR,
    "deepseek_symptom_expression_outputs",
    "deepseek_candidates_review_by_symptom.xlsx"
)

OUTPUT_FILE = os.path.join(
    BASE_DIR,
    "symptom_expression_reference_with_deepseek_candidates.xlsx"
)

DELETE_TXT_M3_FILE = os.path.join(
    BASE_DIR,
    "delete_for_M3_augmented_pure.txt"
)

DELETE_TXT_M4_FILE = os.path.join(
    BASE_DIR,
    "delete_for_M4_augmented_pure.txt"
)

DELETE_TXT_M5_FILE = os.path.join(
    BASE_DIR,
    "delete_for_M5_augmented_pure.txt"
)

LOG_FILE = os.path.join(
    BASE_DIR,
    "deepseek_symptom_expression_outputs",
    "append_candidate_summary_log.xlsx"
)


# 如果 symptom_expression_reference.xlsx 第一行是表头，改成 True
# 如果第一行就是正式症状数据，保持 False
REFERENCE_HAS_HEADER = False


# ==========================================================
# 1. 基础清理函数
# ==========================================================
def clean_text(x):
    if pd.isna(x):
        return ""

    x = str(x)
    x = x.replace("\n", "")
    x = x.replace("\r", "")
    x = x.replace("\t", "")
    x = x.strip()

    return x


def normalize_text(x):
    """
    用于判断重复。
    去掉空白和常见标点。
    """
    x = clean_text(x)
    x = re.sub(r"\s+", "", x)

    for ch in [
        "，", "。", "；", "、", ",", ".", ";",
        "：", ":", "！", "!", "？", "?",
        "（", "）", "(", ")", "[", "]", "【", "】",
        "\"", "'", "“", "”", "‘", "’"
    ]:
        x = x.replace(ch, "")

    return x


# ==========================================================
# 2. 把候选汇总里的 item 变成纯文本 phrase
# ==========================================================
def extract_phrase_from_summary_item(item):
    """
    输入例子：
      入睡困难(n=67, variant, high)
      早醒(n=4, new_candidate, high)

    输出：
      入睡困难
      早醒
    """
    item = clean_text(item)

    if item == "":
        return ""

    # 去掉最后的 (n=..., ...)
    item = re.sub(r"\(n=\d+.*?\)$", "", item).strip()

    return item


def split_candidate_summary(summary_text):
    """
    把一整串：
      入睡困难(n=67, variant, high)；入睡晚(n=6, variant, high)

    变成列表：
      ["入睡困难", "入睡晚"]
    """
    summary_text = clean_text(summary_text)

    if summary_text == "":
        return []

    parts = summary_text.split("；")

    phrases = []

    for part in parts:
        phrase = extract_phrase_from_summary_item(part)

        if phrase != "":
            phrases.append(phrase)

    return phrases


# ==========================================================
# 3. Excel 样式复制
# ==========================================================
def copy_style_from_left(ws, row_idx, target_col):
    """
    新增单元格复制左侧单元格样式。
    """
    if target_col <= 1:
        return

    source = ws.cell(row_idx, target_col - 1)
    target = ws.cell(row_idx, target_col)

    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def first_blank_col_from_e(ws, row_idx):
    """
    找到某一行从 E列开始的第一个空白单元格。
    """
    col_idx = 5

    while True:
        value = clean_text(ws.cell(row_idx, col_idx).value)

        if value == "":
            return col_idx

        col_idx += 1


# ==========================================================
# 4. 删除词典导出函数
# ==========================================================
def collect_phrases_from_rows(ws, start_row, row_keep_func):
    """
    从满足条件的症状行中，收集 E列以后所有表述。
    row_keep_func(ws, r) 返回 True 表示这一行纳入导出。
    """
    phrases = []
    seen = set()
    kept_symptom_rows = []

    for r in range(start_row, ws.max_row + 1):
        symptom = clean_text(ws.cell(r, 1).value)

        if symptom == "":
            continue

        if not row_keep_func(ws, r):
            continue

        kept_symptom_rows.append({
            "Excel行号": r,
            "症状": symptom,
            "C列": clean_text(ws.cell(r, 3).value),
            "D列": clean_text(ws.cell(r, 4).value)
        })

        # E列以后都是该症状对应的表述
        for c in range(5, ws.max_column + 1):
            phrase = clean_text(ws.cell(r, c).value)

            if phrase == "":
                continue

            phrase_norm = normalize_text(phrase)

            if phrase_norm == "":
                continue

            if phrase_norm in seen:
                continue

            seen.add(phrase_norm)
            phrases.append(phrase)

    # 长词优先，后续删词更干净
    phrases = sorted(
        phrases,
        key=lambda x: len(str(x)),
        reverse=True
    )

    return phrases, kept_symptom_rows


def write_txt(phrases, output_file):
    """
    写出一行一个表述的 txt。
    """
    with open(output_file, "w", encoding="utf-8") as f:
        for phrase in phrases:
            f.write(phrase + "\n")


def keep_all_symptom_rows(ws, r):
    """
    M5：所有症状行都保留。
    """
    return True


def keep_m4_rows(ws, r):
    """
    M4：
    C列包含“文本症状”的行。
    """
    c_value = clean_text(ws.cell(r, 3).value)

    if "文本症状" in c_value:
        return True

    return False


def keep_m3_rows(ws, r):
    """
    M3：
    1. C列包含“文本症状”
       +
    2. C列包含“问卷症状” 且 D列为空白

    表格列对应：
      A列：症状名
      B列：复合症状体系/其他信息
      C列：症状来源/类型
      D列：空白 或 4个问卷
      E列以后：症状表述
    """
    c_value = clean_text(ws.cell(r, 3).value)
    d_value = clean_text(ws.cell(r, 4).value)

    # M4 部分：文本症状
    if "文本症状" in c_value:
        return True

    # M3 额外部分：问卷症状，但不属于4个问卷
    if "问卷症状" in c_value and d_value == "":
        return True

    return False


# ==========================================================
# 5. 主程序
# ==========================================================
def main():
    print("REFERENCE_FILE:", REFERENCE_FILE)
    print("CANDIDATE_FILE:", CANDIDATE_FILE)
    print("OUTPUT_FILE:", OUTPUT_FILE)
    print("DELETE_TXT_M3_FILE:", DELETE_TXT_M3_FILE)
    print("DELETE_TXT_M4_FILE:", DELETE_TXT_M4_FILE)
    print("DELETE_TXT_M5_FILE:", DELETE_TXT_M5_FILE)
    print("LOG_FILE:", LOG_FILE)

    if not os.path.exists(REFERENCE_FILE):
        raise FileNotFoundError("找不到原始参考表：" + REFERENCE_FILE)

    if not os.path.exists(CANDIDATE_FILE):
        raise FileNotFoundError("找不到候选整理表：" + CANDIDATE_FILE)

    os.makedirs(os.path.dirname(LOG_FILE), exist_ok=True)

    # ------------------------------------------------------
    # 5.1 读取原始症状表
    # ------------------------------------------------------
    wb = load_workbook(REFERENCE_FILE)
    ws = wb.active

    start_row = 2 if REFERENCE_HAS_HEADER else 1

    # 建立：A列症状 -> 行号
    symptom_to_row = {}

    # 建立：每个症状行已有表述，用于去重
    existing_by_row = {}

    for r in range(start_row, ws.max_row + 1):
        symptom = clean_text(ws.cell(r, 1).value)

        if symptom == "":
            continue

        symptom_norm = normalize_text(symptom)

        if symptom_norm == "":
            continue

        symptom_to_row[symptom_norm] = r

        existing_norms = set()

        # E列以后已有表述
        for c in range(5, ws.max_column + 1):
            value = clean_text(ws.cell(r, c).value)

            if value != "":
                existing_norms.add(normalize_text(value))

        existing_by_row[r] = existing_norms

    print("原始表 A列症状数：", len(symptom_to_row))

    # ------------------------------------------------------
    # 5.2 读取“每个症状候选汇总”sheet
    # ------------------------------------------------------
    candidate_df = pd.read_excel(
        CANDIDATE_FILE,
        sheet_name="每个症状候选汇总"
    )

    required_cols = ["症状", "候选表述汇总"]

    for col in required_cols:
        if col not in candidate_df.columns:
            raise ValueError(
                f"候选表缺少列：{col}。当前列名为：{list(candidate_df.columns)}"
            )

    # ------------------------------------------------------
    # 5.3 逐症状追加
    # ------------------------------------------------------
    added_rows = []
    skipped_rows = []

    total_added = 0

    for _, row in candidate_df.iterrows():
        symptom = clean_text(row["症状"])
        candidate_summary = clean_text(row["候选表述汇总"])

        if symptom == "" or candidate_summary == "":
            continue

        symptom_norm = normalize_text(symptom)

        if symptom_norm not in symptom_to_row:
            skipped_rows.append({
                "症状": symptom,
                "候选表述": "",
                "跳过原因": "原始 reference A列找不到这个症状"
            })
            continue

        row_idx = symptom_to_row[symptom_norm]
        existing_norms = existing_by_row.get(row_idx, set())

        phrases = split_candidate_summary(candidate_summary)

        for phrase in phrases:
            phrase_norm = normalize_text(phrase)

            if phrase_norm == "":
                continue

            if phrase_norm in existing_norms:
                skipped_rows.append({
                    "症状": symptom,
                    "候选表述": phrase,
                    "跳过原因": "原表该症状行已存在"
                })
                continue

            col_idx = first_blank_col_from_e(ws, row_idx)

            ws.cell(row_idx, col_idx).value = phrase
            copy_style_from_left(ws, row_idx, col_idx)

            existing_norms.add(phrase_norm)
            total_added += 1

            added_rows.append({
                "症状": symptom,
                "候选表述": phrase,
                "追加到原表行号": row_idx,
                "追加到列号": col_idx
            })

        existing_by_row[row_idx] = existing_norms

    # ------------------------------------------------------
    # 5.4 保存增强版 reference
    # ------------------------------------------------------
    wb.save(OUTPUT_FILE)

    # ------------------------------------------------------
    # 5.5 导出 M4 删除词典
    # ------------------------------------------------------
    m4_phrases, m4_symptom_rows = collect_phrases_from_rows(
        ws=ws,
        start_row=start_row,
        row_keep_func=keep_m4_rows
    )

    write_txt(
        phrases=m4_phrases,
        output_file=DELETE_TXT_M4_FILE
    )

    # ------------------------------------------------------
    # 5.6 导出 M3 删除词典
    # ------------------------------------------------------
    m3_phrases, m3_symptom_rows = collect_phrases_from_rows(
        ws=ws,
        start_row=start_row,
        row_keep_func=keep_m3_rows
    )

    write_txt(
        phrases=m3_phrases,
        output_file=DELETE_TXT_M3_FILE
    )

    # ------------------------------------------------------
    # 5.7 导出 M5 删除词典
    # ------------------------------------------------------
    m5_phrases, m5_symptom_rows = collect_phrases_from_rows(
        ws=ws,
        start_row=start_row,
        row_keep_func=keep_all_symptom_rows
    )

    write_txt(
        phrases=m5_phrases,
        output_file=DELETE_TXT_M5_FILE
    )

    # ------------------------------------------------------
    # 5.8 输出日志
    # ------------------------------------------------------
    added_df = pd.DataFrame(added_rows)
    skipped_df = pd.DataFrame(skipped_rows)
    m4_rows_df = pd.DataFrame(m4_symptom_rows)
    m3_rows_df = pd.DataFrame(m3_symptom_rows)
    m5_rows_df = pd.DataFrame(m5_symptom_rows)

    with pd.ExcelWriter(LOG_FILE, engine="openpyxl") as writer:
        added_df.to_excel(writer, sheet_name="已追加", index=False)
        skipped_df.to_excel(writer, sheet_name="已跳过", index=False)
        m4_rows_df.to_excel(writer, sheet_name="M4纳入症状行", index=False)
        m3_rows_df.to_excel(writer, sheet_name="M3纳入症状行", index=False)
        m5_rows_df.to_excel(writer, sheet_name="M5纳入症状行", index=False)

    print("\n完成。")
    print("实际新增表述数：", total_added)
    print("M4 纳入症状行数：", len(m4_symptom_rows))
    print("M4 删除词典唯一表述数：", len(m4_phrases))
    print("M3 纳入症状行数：", len(m3_symptom_rows))
    print("M3 删除词典唯一表述数：", len(m3_phrases))
    print("M5 纳入症状行数：", len(m5_symptom_rows))
    print("M5 删除词典唯一表述数：", len(m5_phrases))

    print("\n增强版表格：", OUTPUT_FILE)
    print("追加日志：", LOG_FILE)
    print("M4 删除词典 txt：", DELETE_TXT_M4_FILE)
    print("M3 删除词典 txt：", DELETE_TXT_M3_FILE)
    print("M5 删除词典 txt：", DELETE_TXT_M5_FILE)

    print("\n后续 M4 删词用：")
    print(DELETE_TXT_M4_FILE)

    print("\n后续 M3 删词用：")
    print(DELETE_TXT_M3_FILE)

    print("\n后续 M5 删词用：")
    print(DELETE_TXT_M5_FILE)


if __name__ == "__main__":
    main()